"""PDF to Excel conversion service."""
import os
import logging

logger = logging.getLogger(__name__)


class PdfToExcelError(Exception):
    """Custom exception for PDF to Excel conversion failures."""
    pass


def pdf_to_excel(input_path: str, output_path: str) -> dict:
    """
    Convert a PDF file containing tables to an Excel spreadsheet.

    Args:
        input_path: Path to the input PDF
        output_path: Path for the output Excel file

    Returns:
        dict with total_pages, tables_found, output_size

    Raises:
        PdfToExcelError: If conversion fails
    """
    try:
        import tabula

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # Read all tables from the PDF
        tables = tabula.read_pdf(
            input_path, pages="all", multiple_tables=True, silent=True
        )

        if not tables:
            raise PdfToExcelError(
                "No tables found in the PDF. This tool works best with PDFs that contain tabular data."
            )

        # Write tables to Excel, each table on its own sheet
        import openpyxl

        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        for idx, df in enumerate(tables, 1):
            sheet_name = f"Table_{idx}"
            ws = wb.create_sheet(title=sheet_name)

            # Write header
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=str(col_name))

            # Write data
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    cell_value = value
                    # Convert NaN to empty string
                    if isinstance(value, float) and str(value) == "nan":
                        cell_value = ""
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)

        wb.save(output_path)

        output_size = os.path.getsize(output_path)

        logger.info(
            f"PDF→Excel: {len(tables)} tables extracted → {output_size} bytes"
        )

        return {
            "tables_found": len(tables),
            "output_size": output_size,
        }

    except PdfToExcelError:
        raise
    except ImportError as e:
        raise PdfToExcelError(f"Required library not installed: {e}")
    except Exception as e:
        raise PdfToExcelError(f"Failed to convert PDF to Excel: {str(e)}")
